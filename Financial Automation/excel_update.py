import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from config import REPORTS_DIR

def update_excel_files(cleaned_data):
    os.makedirs(REPORTS_DIR, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    title_font = Font(bold=True, size=16, name='Calibri')
    header_font = Font(bold=True, size=12, name='Calibri')
    data_font = Font(size=11, name='Calibri')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    total_fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
    sub_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    thin = Side(style='thin')

    border_style = Border(left=thin, right=thin, top=thin, bottom=thin)

    for symbol, data in cleaned_data.items():
        for stmt_key, df in data.items():
            if stmt_key != 'prices':
                ws = wb.create_sheet(title=f'{symbol}_{stmt_key[:10]}')

                ws.merge_cells('A1:J1')
                title_cell = ws['A1']
                title_cell.value = f'{symbol} {stmt_key.replace("_", " ").title()}'
                title_cell.font = title_font
                title_cell.fill = PatternFill(start_color='365A76', end_color='365A76', fill_type='solid')
                title_cell.alignment = center_align
                ws.row_dimensions[1].height = 30

                rows = dataframe_to_rows(df, index=True, header=True)
                for r_idx, row in enumerate(rows, 1):
                    for c_idx, value in enumerate(row, 1):
                        r = r_idx + 1
                        c = c_idx
                        cell = ws.cell(row=r, column=c, value=value)
                        cell.border = border_style

                num_cols = df.shape[1] + 1
                num_rows = df.shape[0] + 1

                for c in range(2, num_cols + 1):
                    cell = ws.cell(row=2, column=c)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align

                for r in range(3, num_rows + 2):
                    cell = ws.cell(row=r, column=1)
                    cell.font = data_font
                    cell.alignment = left_align

                for r in range(3, num_rows + 2):
                    for c in range(2, num_cols + 1):
                        cell = ws.cell(row=r, column=c)
                        cell.font = data_font
                        cell.alignment = right_align
                        cell.number_format = '#,##0.00'

                total_patterns = ['Total', 'Net', 'Operating']
                for r in range(3, num_rows + 2):
                    label_cell = ws.cell(row=r, column=1)
                    label = str(label_cell.value).upper()
                    if any(pat in label for pat in total_patterns):
                        label_cell.fill = sub_fill
                        label_cell.font = Font(bold=True)
                        for c in range(2, num_cols + 1):
                            val_cell = ws.cell(row=r, column=c)
                            val_cell.fill = total_fill
                            val_cell.font = Font(bold=True)

                ws.freeze_panes = 'B3'
                ws.sheet_view.showGridLines = False

                dims = {}
                for row in ws.rows:
                    for cell in row:
                        if cell.value:
                            col_letter = cell.column_letter
                            dims[col_letter] = max(dims.get(col_letter, 0), len(str(cell.value)))
                for col, value in dims.items():
                    ws.column_dimensions[col].width = min(value + 2, 18)

                for r in range(1, ws.max_row + 1):
                    ws.row_dimensions[r].height = 20

    wb.save(f'{REPORTS_DIR}/financial_statements_pro.xlsx')
    print('Premium Excel financial statements ready!')

